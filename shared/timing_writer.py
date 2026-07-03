"""
shared/timing_writer.py — Export Excel des mesures de performance.

Structure : 3 niveaux d'en-tête regroupés par microservice :
  Ligne 1 : Microservice  (fusionné sur toutes ses colonnes)
  Ligne 2 : Sous-groupe   (fusionné sur les colonnes du sous-groupe)
             OU libellé   (si pas de sous-groupe → lignes 2+3 fusionnées)
  Ligne 3 : Nom de colonne (pour les colonnes avec sous-groupe)

Un fichier .xlsx par mode :
  • AUTONOMOUS → une ligne par cycle d'orchestration
  • ENHANCED   → une ligne par intention (écrite sur le cycle qui exécute la migration)

Constructeurs : TimingWriter.for_autonomous(path) / TimingWriter.for_enhanced(path)
"""

import asyncio
import logging
import threading
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.timing import PARALLEL_BRANCHES, PARALLEL_STEPS, PIPELINE_ORDER, SEQUENTIAL_CHAINS

logger = logging.getLogger("TimingWriter")

# ── Identifiants de microservices (Ligne 1 des en-têtes) ──────────────────────
_S_HUB = "Hub — Orchestrateur"
_S_COL = "Collector"
_S_DB  = "Database"
_S_HL  = "History Loader"
_S_ML  = "ML Predictor"
_S_MM  = "Metrics Manager"
_S_DI  = "Decision Intelligence"
_S_IM  = "Intent Manager"          # mode enhanced uniquement

# ── Identifiants de sous-groupes (Ligne 2 quand il y a un 3e niveau) ──────────
_SG_PVM    = "Collecte par VM (∥)"
_SG_MI     = "MI — Information Mutuelle"
_SG_DISUB  = "Sous-étapes internes"
_SG_TOPSIS = "TOPSIS — calcul"

# ── Tag spécial pour la colonne SOMME ─────────────────────────────────────────
_SUM = "__SUM__"

# ── Format des colonnes : (clé, microservice, sous-groupe, libellé, type) ─────
#   microservice = None → colonne autonome (lignes 1-3 fusionnées)
#   microservice = _SUM → colonne SOMME   (lignes 1-3 fusionnées, couleur spéciale)
#   sous-groupe  = None → pas de 3e niveau ; lignes 2+3 fusionnées, libellé en ligne 2
#   sous-groupe  ≠ None → 3 niveaux : ligne1=service, ligne2=sg, ligne3=libellé

_AUTONOMOUS_COLS: List[Tuple] = [
    # ── Info (colonnes autonomes) ─────────────────────────────────────────
    ("cycle",     None, None, "Cycle",              "int"),
    ("timestamp", None, None, "Horodatage (UTC)",   "text"),
    ("vm",        None, None, "VM active (source)", "text"),
    ("decision",  None, None, "Décision",           "text"),
    ("to_vm",     None, None, "VM cible",           "text"),
    # ═══ Ordre d'exécution du pipeline (cf. shared.timing.PIPELINE_ORDER) ══
    # ── [Étape 1 ∥] Metrics Manager — SLOs / MI (parallèle avec Collector) ─
    ("mi_compute", _S_MM, _SG_MI, "Calcul MI k-NN", "ms"),
    ("mi_slos",    _S_MM, _SG_MI, "Sélection SLOs", "ms"),
    ("slos_mm",    _S_MM, None,   "Total MM (ms)",  "ms"),
    # ── [Étape 1 ∥] Collector — collecte CPU/RAM (parallèle avec MM) ───────
    ("collect_edge1",  _S_COL, _SG_PVM, "edge1",       "ms"),
    ("collect_edge2",  _S_COL, _SG_PVM, "edge2",       "ms"),
    ("collect_cloud1", _S_COL, _SG_PVM, "cloud1",      "ms"),
    ("collect_cloud2", _S_COL, _SG_PVM, "cloud2",      "ms"),
    ("collect_max",      _S_COL, _SG_PVM, "MAX (lente)",          "ms"),
    ("collect",          _S_COL, None,   "Total (ms)",           "ms"),
    ("collect_overhead", _S_COL, None,   "Overhead réseau (ms)", "ms"),
    # ── [Étape 2] Database — persistance SLOs + métriques (+ décision) ─────
    ("persist_slos",    _S_DB, None, "Store SLOs (ms)",       "ms"),
    ("persist_metrics", _S_DB, None, "Store Métriques ∥ (ms)","ms"),
    ("store_decision",  _S_DB, None, "Store Décision (ms)",   "ms"),
    ("db_total",        _S_DB, None, "Total DB (ms)",         "ms"),
    # ── [Étapes 1 & 3] History Loader — hist. VM active puis 4 VMs ─────────
    ("slos_load_hist", _S_HL, None, "Hist. VM active (ms)", "ms"),
    ("load_histories", _S_HL, None, "Hist. 4 VMs ∥ (ms)",  "ms"),
    ("hl_total",       _S_HL, None, "Total HL (ms)",        "ms"),
    # ── [Étape 4] ML Predictor — prédiction 4 VMs ─────────────────────────
    ("prediction", _S_ML, None, "Prédiction 4 VMs ∥ (ms)", "ms"),
    # ── [Étape 6] Decision Intelligence — détection viol. + TOPSIS ────────
    ("violation_detection", _S_DI, _SG_DISUB,  "→ Détect. viol.",   "ms"),
    ("candidate_filter",    _S_DI, _SG_DISUB,  "→ Filtrage cand.",  "ms"),
    ("topsis_total",        _S_DI, _SG_TOPSIS, "Total TOPSIS",      "ms"),
    ("topsis_matrix",       _S_DI, _SG_TOPSIS, "1. Matrice",        "ms"),
    ("topsis_norm",         _S_DI, _SG_TOPSIS, "2. Normalisation",  "ms"),
    ("topsis_weight",       _S_DI, _SG_TOPSIS, "3. Pondération",    "ms"),
    ("topsis_dist",         _S_DI, _SG_TOPSIS, "4. Distances",      "ms"),
    ("decide_call",  _S_DI, None, "Total DI (ms)",          "ms"),
    ("di_overhead",  _S_DI, None, "Overhead HTTP+logs (ms)", "ms"),
    # ── [Étapes 5 & 7] Hub — détect. viol. (log) + migration (action) ─────
    ("check_violations", _S_HUB, None, "Détect. viol. (ms)", "ms"),
    ("migration",        _S_HUB, None, "Migration (ms)",      "ms"),
    ("hub_total",        _S_HUB, None, "Total Hub (ms)",      "ms"),
    # ── Résumé ────────────────────────────────────────────────────────────
    ("somme", _SUM, None, "SOMME (ms)",       "ms"),
    ("total", None, None, "TOTAL cycle (ms)", "ms"),
]

_ENHANCED_COLS: List[Tuple] = [
    # ── Info (colonnes autonomes) ─────────────────────────────────────────
    ("intent_id",  None, None, "Intent ID",             "text"),
    ("timestamp",  None, None, "Horodatage (UTC)",       "text"),
    ("intention",  None, None, "Intention",              "text"),
    ("vm",         None, None, "VM active (source)",    "text"),
    ("decision",   None, None, "Décision",              "text"),
    ("to_vm",      None, None, "VM cible",              "text"),
    # ═══ Ordre d'exécution du pipeline (cf. shared.timing.PIPELINE_ORDER) ══
    # ── [Étape 0] Intent Manager — réception + LLM (AVANT le cycle) ───────
    ("intent_reception", _S_IM, None, "Réception + LLM (ms)", "ms"),
    # ── [Étape 1 ∥] Metrics Manager — SLOs / MI (parallèle avec Collector) ─
    ("mi_compute", _S_MM, _SG_MI, "Calcul MI k-NN", "ms"),
    ("mi_slos",    _S_MM, _SG_MI, "Sélection SLOs", "ms"),
    ("slos_mm",    _S_MM, None,   "Total MM (ms)",  "ms"),
    # ── [Étape 1 ∥] Collector — collecte CPU/RAM (parallèle avec MM) ───────
    ("collect_edge1",  _S_COL, _SG_PVM, "edge1",       "ms"),
    ("collect_edge2",  _S_COL, _SG_PVM, "edge2",       "ms"),
    ("collect_cloud1", _S_COL, _SG_PVM, "cloud1",      "ms"),
    ("collect_cloud2", _S_COL, _SG_PVM, "cloud2",      "ms"),
    ("collect_max",      _S_COL, _SG_PVM, "MAX (lente)",          "ms"),
    ("collect",          _S_COL, None,   "Total (ms)",           "ms"),
    ("collect_overhead", _S_COL, None,   "Overhead réseau (ms)", "ms"),
    # ── [Étape 2] Database — persistance SLOs + métriques (+ décision) ─────
    ("persist_slos",    _S_DB, None, "Store SLOs (ms)",       "ms"),
    ("persist_metrics", _S_DB, None, "Store Métriques ∥ (ms)","ms"),
    ("store_decision",  _S_DB, None, "Store Décision (ms)",   "ms"),
    ("db_total",        _S_DB, None, "Total DB (ms)",         "ms"),
    # ── [Étapes 1 & 3] History Loader — hist. VM active puis 4 VMs ─────────
    ("slos_load_hist", _S_HL, None, "Hist. VM active (ms)", "ms"),
    ("load_histories", _S_HL, None, "Hist. 4 VMs ∥ (ms)",  "ms"),
    ("hl_total",       _S_HL, None, "Total HL (ms)",        "ms"),
    # ── [Étape 4] ML Predictor — prédiction 4 VMs ─────────────────────────
    ("prediction", _S_ML, None, "Prédiction 4 VMs ∥ (ms)", "ms"),
    # ── [Étape 6] Decision Intelligence — détection viol. + TOPSIS ────────
    ("violation_detection", _S_DI, _SG_DISUB,  "→ Détect. viol.",   "ms"),
    ("candidate_filter",    _S_DI, _SG_DISUB,  "→ Filtrage cand.",  "ms"),
    ("topsis_total",        _S_DI, _SG_TOPSIS, "Total TOPSIS",      "ms"),
    ("topsis_matrix",       _S_DI, _SG_TOPSIS, "1. Matrice",        "ms"),
    ("topsis_norm",         _S_DI, _SG_TOPSIS, "2. Normalisation",  "ms"),
    ("topsis_weight",       _S_DI, _SG_TOPSIS, "3. Pondération",    "ms"),
    ("topsis_dist",         _S_DI, _SG_TOPSIS, "4. Distances",      "ms"),
    ("decide_call",  _S_DI, None, "Total DI (ms)",          "ms"),
    ("di_overhead",  _S_DI, None, "Overhead HTTP+logs (ms)", "ms"),
    # ── [Étapes 5 & 7] Hub — détect. viol. (log) + migration (action) ─────
    ("check_violations", _S_HUB, None, "Détect. viol. (ms)", "ms"),
    ("migration",        _S_HUB, None, "Migration (ms)",      "ms"),
    ("hub_total",        _S_HUB, None, "Total Hub (ms)",      "ms"),
    # ── Résumé ────────────────────────────────────────────────────────────
    ("somme",          _SUM, None, "SOMME (ms)",                               "ms"),
    ("cycle_total",    None, None, "Total cycle (ms)",                         "ms"),
    ("intention_total",None, None, "TOTAL intention : réception→migration (ms)","ms"),
]

_SHEET_AUTONOMOUS = "Autonomous (par cycle)"
_SHEET_ENHANCED   = "Enhanced (par intention)"
_SHEET_LEGEND     = "Légende"

_DEF_AUTONOMOUS = ("durée d'un cycle d'orchestration complet "
                   "(collecte → prédiction → décision → migration).")
_DEF_ENHANCED   = ("de la réception de l'intention (LLM) JUSQU'À l'exécution de la migration. "
                   "TOTAL intention = Réception intention + cycle qui exécute la migration.")

_HEADER_ROWS   = 3
_TRIM_FRACTION = 0.20

# ── Couleurs par microservice ──────────────────────────────────────────────────
_P = PatternFill  # raccourci

# Ligne 1 : teinte sombre  (nom du microservice)
_SVC_DARK = {
    _S_HUB: _P("solid", fgColor="1F3864"),   # bleu marine très sombre
    _S_COL: _P("solid", fgColor="145A32"),   # vert très sombre
    _S_DB:  _P("solid", fgColor="641E16"),   # rouge très sombre
    _S_HL:  _P("solid", fgColor="1A2980"),   # indigo très sombre
    _S_ML:  _P("solid", fgColor="0B5345"),   # sarcelle très sombre
    _S_MM:  _P("solid", fgColor="6E2F1A"),   # brun-orange très sombre
    _S_DI:  _P("solid", fgColor="4A235A"),   # violet très sombre
    _S_IM:  _P("solid", fgColor="7B241C"),   # rouge sombre (enhanced)
}

# Lignes 2+3 : teinte moyenne  (sous-groupe ou libellé direct)
_SVC_MED = {
    _S_HUB: _P("solid", fgColor="2E5496"),   # bleu marine
    _S_COL: _P("solid", fgColor="1E8449"),   # vert
    _S_DB:  _P("solid", fgColor="922B21"),   # rouge
    _S_HL:  _P("solid", fgColor="2471A3"),   # bleu moyen
    _S_ML:  _P("solid", fgColor="0E6655"),   # sarcelle
    _S_MM:  _P("solid", fgColor="935116"),   # brun-orange
    _S_DI:  _P("solid", fgColor="7D3C98"),   # violet
    _S_IM:  _P("solid", fgColor="B03A2E"),   # rouge moyen (enhanced)
}

# Sous-groupes spéciaux (légèrement différents de la teinte "med" du service)
_SG_FILLS = {
    _SG_PVM:    _P("solid", fgColor="216A3E"),  # vert plus riche (Collector)
    _SG_MI:     _P("solid", fgColor="A04000"),  # orange sombre (Metrics Manager)
    _SG_DISUB:  _P("solid", fgColor="884EA0"),  # violet clair (DI sous-étapes)
    _SG_TOPSIS: _P("solid", fgColor="6C3483"),  # violet sombre (DI TOPSIS)
}

_FILL_HEAD = _P("solid", fgColor="1F3864")   # en-têtes info/total
_FILL_SUM  = _P("solid", fgColor="7B3F00")   # SOMME (brun/orange)
_FILL_GRP  = _P("solid", fgColor="2E5496")   # fallback

_FONT_HEAD  = Font(bold=True, color="FFFFFF", size=10)
_FONT_TITLE = Font(bold=True, color="1F3864", size=13)
_FONT_SUB   = Font(bold=True, color="1F3864", size=11)
_ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_R    = Alignment(horizontal="right",  vertical="center")
_ALIGN_C    = Alignment(horizontal="center", vertical="center")
_ALIGN_L    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_THIN       = Side(style="thin", color="9DB0CE")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


class TimingWriter:
    """Persiste les mesures de performance d'un mode dans un .xlsx présentable."""

    def __init__(
        self,
        path: str,
        columns: List[Tuple],
        sheet_name: str,
        total_definition: str,
        max_bytes: int = 100 * 1024 * 1024,
    ) -> None:
        self.path             = Path(path)
        self.columns          = columns
        self.sheet_name       = sheet_name
        self.total_definition = total_definition
        self.max_bytes        = max_bytes
        self._lock            = threading.Lock()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if not self.path.exists():
            self._create_workbook()
        elif not self._schema_matches():
            # L'ordre / la liste des colonnes a changé depuis la création du
            # fichier : l'ancien en-tête serait désaligné avec les nouvelles
            # lignes. On recrée (les mesures sont des logs régénérables).
            logger.info(f"♻️  Schéma de colonnes modifié — recréation de {self.path}")
            self.path.unlink(missing_ok=True)
            self._create_workbook()

    # ── Signature de schéma (détecte tout changement d'ordre/de colonnes) ─────

    def _schema_signature(self) -> str:
        """Signature stable = liste ordonnée des clés de colonnes."""
        return ";".join(str(c[0]) for c in self.columns)

    def _schema_matches(self) -> bool:
        """True si le fichier existant a été créé avec les mêmes colonnes."""
        try:
            wb  = load_workbook(self.path, read_only=True)
            sig = wb.properties.keywords or ""
            wb.close()
            return sig == self._schema_signature()
        except Exception:
            return False

    # ── Constructeurs dédiés ──────────────────────────────────────────────────

    @classmethod
    def for_autonomous(cls, path: str, max_bytes: int = 100 * 1024 * 1024) -> "TimingWriter":
        return cls(path, _AUTONOMOUS_COLS, _SHEET_AUTONOMOUS, _DEF_AUTONOMOUS, max_bytes)

    @classmethod
    def for_enhanced(cls, path: str, max_bytes: int = 100 * 1024 * 1024) -> "TimingWriter":
        return cls(path, _ENHANCED_COLS, _SHEET_ENHANCED, _DEF_ENHANCED, max_bytes)

    # ── Création / mise en forme ──────────────────────────────────────────────

    def _create_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        self._build_sheet(ws, self.columns)
        self._build_legend(wb.create_sheet(_SHEET_LEGEND))
        wb.properties.keywords = self._schema_signature()   # signature de schéma
        wb.save(self.path)
        logger.info(f"📊 Fichier de mesures créé : {self.path}")

    def _build_sheet(self, ws, cols: List[Tuple]) -> None:
        n = len(cols)

        # ── PASSE 1 : Ligne 1 — noms des microservices ───────────────────────
        col, i = 1, 0
        while i < n:
            _, svc, _, _, _ = cols[i]
            if svc is None or svc == _SUM:
                # Colonne autonome → traitée en passe 3
                col += 1; i += 1
                continue
            # Trouver le span complet de ce microservice
            j = i
            while j < n and cols[j][1] == svc:
                j += 1
            span = j - i
            if span > 1:
                ws.merge_cells(
                    start_row=1, start_column=col,
                    end_row=1,   end_column=col + span - 1
                )
            cell = ws.cell(row=1, column=col, value=svc)
            self._style_head(cell, _SVC_DARK.get(svc, _FILL_HEAD))
            for cc in range(col, col + span):
                ws.cell(row=1, column=cc).border = _BORDER
            col += span; i = j

        # ── PASSE 2 : Ligne 2 — sous-groupes OU libellés (si pas de sg) ──────
        col, i = 1, 0
        while i < n:
            _, svc, sg, label, _ = cols[i]
            if svc is None or svc == _SUM:
                col += 1; i += 1
                continue
            if sg is None:
                # Pas de sous-groupe : fusionner lignes 2+3 et écrire le libellé
                ws.merge_cells(
                    start_row=2, start_column=col,
                    end_row=3,   end_column=col
                )
                cell = ws.cell(row=2, column=col, value=label)
                self._style_head(cell, _SVC_MED.get(svc, _FILL_GRP))
                ws.cell(row=3, column=col).border = _BORDER
                col += 1; i += 1
            else:
                # Trouver le span du sous-groupe (même service ET même sg)
                j = i
                while j < n and cols[j][1] == svc and cols[j][2] == sg:
                    j += 1
                span = j - i
                if span > 1:
                    ws.merge_cells(
                        start_row=2, start_column=col,
                        end_row=2,   end_column=col + span - 1
                    )
                cell = ws.cell(row=2, column=col, value=sg)
                fill = _SG_FILLS.get(sg, _SVC_MED.get(svc, _FILL_GRP))
                self._style_head(cell, fill)
                for cc in range(col, col + span):
                    ws.cell(row=2, column=cc).border = _BORDER
                col += span; i = j

        # ── PASSE 3 : Ligne 3 — libellés des colonnes + fusions autonomes ────
        col = 1
        for _, svc, sg, label, _ in cols:
            if svc is None:
                # Colonne info : fusionner lignes 1-3
                ws.merge_cells(
                    start_row=1, start_column=col,
                    end_row=3,   end_column=col
                )
                cell = ws.cell(row=1, column=col, value=label)
                self._style_head(cell, _FILL_HEAD)
                ws.cell(row=2, column=col).border = _BORDER
                ws.cell(row=3, column=col).border = _BORDER
            elif svc == _SUM:
                # Colonne SOMME : fusionner lignes 1-3, couleur spéciale
                ws.merge_cells(
                    start_row=1, start_column=col,
                    end_row=3,   end_column=col
                )
                cell = ws.cell(row=1, column=col, value=label)
                self._style_head(cell, _FILL_SUM)
                ws.cell(row=2, column=col).border = _BORDER
                ws.cell(row=3, column=col).border = _BORDER
            elif sg is None:
                pass  # déjà géré en passe 2 (lignes 2+3 fusionnées)
            else:
                # Colonne avec sous-groupe : libellé en ligne 3
                cell = ws.cell(row=3, column=col, value=label)
                fill = _SG_FILLS.get(sg, _SVC_MED.get(svc, _FILL_GRP))
                self._style_head(cell, fill)
            col += 1

        # ── Dimensions ────────────────────────────────────────────────────────
        for idx, (_, _, _, lbl, _) in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(10, min(28, len(lbl) + 2))
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24
        ws.row_dimensions[3].height = 30
        ws.freeze_panes = "A4"

    def _build_legend(self, ws) -> None:
        lines: List[Tuple] = [
            ("Légende — Mesures de performance du pipeline", _FONT_TITLE),
            ("", None),
            ("Toutes les durées sont exprimées en millisecondes (ms).", None),
            (f"Définition du TOTAL : {self.total_definition}", None),
            ("", None),
            ("Structure du tableau — 3 niveaux d'en-tête :", _FONT_SUB),
            ("  Ligne 1 (teinte sombre)  : Microservice responsable de l'étape.", None),
            ("  Ligne 2 (teinte moyenne) : Sous-groupe fonctionnel au sein du service.", None),
            ("                             Si pas de sous-groupe, lignes 2+3 fusionnées avec le libellé.", None),
            ("  Ligne 3 (teinte moyenne) : Nom individuel de la colonne.", None),
            ("", None),
            ("Ordre des colonnes = ordre d'exécution du pipeline (hub _run_flow) :", _FONT_SUB),
        ]
        for _step in PIPELINE_ORDER:
            lines.append((f"     {_step}", None))
        lines += [
            ("  ∥ = étapes lancées en parallèle (durée = branche la plus lente, pas la somme).", None),
            ("  check_violations (Hub) s'exécute juste avant la décision DI ; le bloc Hub est", None),
            ("  regroupé en fin car son coût dominant est la migration (action finale).", None),
            ("", None),
            ("═" * 80, None),
            ("SOMME (colonne 'SOMME (ms)') — quelles colonnes sommer pour l'obtenir :", _FONT_SUB),
            ("  check_violations + migration                          (Hub)", None),
            ("  + collect                                              (Collector — total, pas la somme des VMs)", None),
            ("  + persist_slos + persist_metrics + store_decision      (Database)", None),
            ("  + slos_load_hist + load_histories                      (History Loader)", None),
            ("  + prediction                                           (ML Predictor)", None),
            ("  + slos_mm                                              (Metrics Manager)", None),
            ("  + decide_call                                          (Decision Intelligence — total, pas les sous-étapes)", None),
            ("  EXCLUES de la SOMME (ce sont des sous-parties déjà comptées dans leur total parent) :", None),
            ("     mi_compute, mi_slos (inclus dans slos_mm) ; violation_detection, candidate_filter,", None),
            ("     topsis_total, topsis_matrix/norm/weight/dist (inclus dans decide_call) ;", None),
            ("     collect_edge1..cloud2, collect_max (inclus dans collect) ; collect_overhead, di_overhead (dérivées).", None),
            ("  TOTAL (colonne 'TOTAL cycle (ms)') > SOMME, car TOTAL est mesuré directement par le hub", None),
            ("  (chrono autour de tout le cycle) et inclut l'overhead HTTP/asyncio ENTRE les étapes,", None),
            ("  que la SOMME (addition des durées internes) ne peut pas capturer.", None),
            ("", None),
            ("═" * 80, None),
            ("PARALLÉLISME — 2 niveaux distincts dans le pipeline :", _FONT_SUB),
            ("", None),
            ("Niveau 1 — ENTRE BRANCHES (deux services différents, un seul cycle) :", None),
        ]
        for b in PARALLEL_BRANCHES:
            lines.append((f"     Branche A : {' → '.join(b['branch_a'])}", None))
            lines.append((f"                 ({b['branch_a_label']})", None))
            lines.append((f"     Branche B : {' → '.join(b['branch_b'])}", None))
            lines.append((f"                 ({b['branch_b_label']})", None))
            lines.append((f"     A ∥ B  →  {b['note']}", None))
        lines += [
            ("     Durée du bloc = max(Branche A, Branche B) — visible dans les logs console", None),
            ("     (\"slos_and_collect_parallel\"), PAS persistée comme colonne Excel séparée.", None),
            ("", None),
            ("Niveau 2 — INTERNE à une étape (4 VMs simultanées via asyncio.gather) :", None),
        ]
        for name, desc in PARALLEL_STEPS.items():
            lines.append((f"     •  {name}  —  {desc}", None))
        lines += [
            ("", None),
            ("La durée d'une étape parallèle (niveau 1 ou 2) = la branche/VM la plus LENTE, jamais la somme.", None),
            ("Ne jamais multiplier une valeur parallèle par 4 (ou par 2) pour obtenir un temps total.", None),
            ("", None),
            ("═" * 80, None),
            ("SÉQUENTIEL — aucun parallélisme, chaque étape attend la sortie de la précédente :", _FONT_SUB),
            ("  (complément du bloc PARALLÉLISME ci-dessus — TOUTES les autres colonnes du tableau,", None),
            ("  hors les 2 niveaux listés plus haut, s'exécutent l'une après l'autre.)", None),
            ("", None),
        ]
        for i, sc in enumerate(SEQUENTIAL_CHAINS, start=1):
            lines.append((f"  {i}. " + " → ".join(sc["chain"]), None))
            lines.append((f"     {sc['note']}", None))
        lines += [
            ("", None),
            ("Récapitulatif pour lire une ligne du tableau de gauche à droite :", None),
            ("  Metrics Manager ∥ Collector  →  Database  →  History Loader  →  ML Predictor  →", None),
            ("  Decision Intelligence  →  Hub (migration). Seul le premier bloc est parallèle (niveau 1) ;", None),
            ("  tout le reste s'enchaîne séquentiellement, microservice après microservice.", None),
            ("", None),
            ("Collecte par VM :", _FONT_SUB),
            ("  edge1, edge2, cloud1, cloud2 : temps de réponse individuel mesuré dans le Collector.", None),
            ("  MAX (lente) = max(edge1, edge2, cloud1, cloud2) = la VM qui a dominé la collecte.", None),
            ("  Total (ms)  = appel HTTP complet hub→collector→hub (inclut transport + FastAPI + sérialisation).", None),
            ("  Overhead réseau = Total − MAX  =  transport HTTP×2 + FastAPI + asyncio.gather + JSON sérialis.", None),
            ("  (valeur dérivée, NON incluse dans SOMME)", None),
            ("", None),
            ("Decision Intelligence — sous-étapes :", _FONT_SUB),
            ("  → Détect. viol.   : prédictions vs seuils SLO sur la VM active (violation_detector).", None),
            ("  → Filtrage cand.  : VMs pré-satisfaisant les SLOs (filtrage avant TOPSIS).", None),
            ("  Total TOPSIS      : temps total du calcul TOPSIS (matrice + norm + poids + distances).", None),
            ("  Total DI (ms)     : appel HTTP complet hub→/decide (= transport + toutes sous-étapes + logs).", None),
            ("  Overhead HTTP+logs = Total DI − (Détect. + Filtrage + TOPSIS)", None),
            ("                     = transport HTTP×2 + FastAPI + logs Python dans decide() + construction retour.", None),
            ("  (valeur dérivée, NON incluse dans SOMME)", None),
            ("", None),
            ("Metrics Manager — slos_mm :", _FONT_SUB),
            ("  MI + SLOs : calcul de l'Information Mutuelle + mise à jour des SLOs secondaires.", None),
            ("  (appel hub→metrics_manager /compute ou /validate)", None),
            ("", None),
            ("History Loader :", _FONT_SUB),
            ("  Hist. VM active : historique de la VM du service (step 1, pour le calcul MI).", None),
            ("  Hist. 4 VMs ∥   : historiques des 4 VMs en parallèle (step 6, pour la prédiction).", None),
            ("", None),
            ("Database :", _FONT_SUB),
            ("  Store SLOs      : écriture des SLOs (primaire + secondaires MI) en base.", None),
            ("  Store Métriques∥: écriture des métriques des 4 VMs en parallèle.", None),
            ("  Store Décision  : écriture de la décision en base (uniquement si migration).", None),
        ]
        for r, (text, font) in enumerate(lines, start=1):
            cell = ws.cell(row=r, column=1, value=text)
            cell.alignment = _ALIGN_L
            if font is not None:
                cell.font = font
        ws.column_dimensions["A"].width = 120

    @staticmethod
    def _style_head(cell, fill: PatternFill) -> None:
        cell.fill      = fill
        cell.font      = _FONT_HEAD
        cell.alignment = _ALIGN_CTR
        cell.border    = _BORDER

    # ── Écriture ──────────────────────────────────────────────────────────────

    def _row_values(self, data: Dict[str, Any]) -> List[Any]:
        vals: List[Any] = []
        for key, _svc, _sg, _lbl, kind in self.columns:
            v = data.get(key)
            if v is None:
                vals.append(None)
            elif kind == "ms":
                try:
                    vals.append(round(float(v), 3))
                except (TypeError, ValueError):
                    vals.append(None)
            elif kind == "int":
                try:
                    vals.append(int(v))
                except (TypeError, ValueError):
                    vals.append(v)
            else:
                vals.append(str(v))
        return vals

    def _append(self, data: Dict[str, Any]) -> None:
        with self._lock:
            try:
                wb = self._load_safe()
                ws = wb[self.sheet_name]
                ws.append(self._row_values(data))
                r = ws.max_row
                for idx, (_key, _svc, _sg, _lbl, kind) in enumerate(self.columns, start=1):
                    cell        = ws.cell(row=r, column=idx)
                    cell.border = _BORDER
                    if kind == "ms":
                        cell.number_format = "0.00"
                        cell.alignment     = _ALIGN_R
                    elif kind == "int":
                        cell.number_format = "0"
                        cell.alignment     = _ALIGN_C
                    else:
                        cell.alignment     = _ALIGN_L
                wb.save(self.path)
                self._maybe_trim()
            except Exception as exc:
                logger.warning(
                    f"⚠️  TimingWriter — échec écriture [{self.sheet_name}] : {exc}"
                )

    def _load_safe(self) -> Any:
        try:
            return load_workbook(self.path)
        except Exception:
            self._create_workbook()
            return load_workbook(self.path)

    def _maybe_trim(self) -> None:
        try:
            size = self.path.stat().st_size
            if size <= self.max_bytes:
                return
            wb = load_workbook(self.path)
            ws = wb[self.sheet_name]
            total = ws.max_row - _HEADER_ROWS
            if total <= 0:
                return
            to_del = max(1, int(total * _TRIM_FRACTION))
            for _ in range(to_del):
                ws.delete_rows(_HEADER_ROWS + 1)
            wb.save(self.path)
            logger.info(
                f"🗑️  TimingWriter — rognage {to_del} lignes "
                f"(taille {size // 1024} Ko > {self.max_bytes // 1024} Ko)"
            )
        except Exception as exc:
            logger.debug(f"TimingWriter._maybe_trim: {exc}")

    async def write(self, data: Dict[str, Any]) -> None:
        await asyncio.to_thread(self._append, data)
