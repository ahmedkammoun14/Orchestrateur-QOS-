import asyncio
import logging
import threading
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

logger = logging.getLogger("ExcelWriter")

_SHEETS: Dict[str, List[str]] = {
    "Métriques":      ["timestamp", "vm_id", "latency", "cpu_usage", "ram_usage", "reliability"],
    "Décisions":      ["timestamp", "decision", "from_vm", "to_vm", "reason", "cycle"],
    "SLOs":           ["timestamp", "metric", "threshold", "operator", "weight", "is_primary"],
    "Intentions_LLM": ["timestamp", "intention", "nb_slos"],
}

_TRIM_FRACTION = 0.20  # retire 20 % des lignes les plus anciennes lors d'une rotation


class ExcelWriter:
    """
    Écrit les données de l'orchestrateur dans un fichier .xlsx.
    Rotation automatique quand la taille dépasse max_bytes :
    les 20 % de lignes les plus anciennes de chaque feuille sont supprimées.
    Toutes les écritures disque sont déléguées à un thread via asyncio.to_thread
    pour ne pas bloquer la boucle asyncio.
    """

    def __init__(self, path: str, max_bytes: int) -> None:
        self.path      = Path(path)
        self.max_bytes = max_bytes
        self._lock     = threading.Lock()
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if not self.path.exists():
            self._create_workbook()

    # ── Init ──────────────────────────────────────────────────────────────

    def _create_workbook(self) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        for name, headers in _SHEETS.items():
            ws = wb.create_sheet(name)
            ws.append(headers)
        wb.save(self.path)
        logger.info(f"📊 Fichier Excel créé : {self.path}")

    # ── Core write ────────────────────────────────────────────────────────

    def _load_workbook_safe(self) -> Workbook:
        """Charge le workbook ; recrée le fichier s'il est corrompu."""
        try:
            return load_workbook(self.path)
        except Exception:
            logger.warning(
                f"⚠️  ExcelWriter — fichier corrompu détecté, recréation : {self.path}"
            )
            self._create_workbook()
            return load_workbook(self.path)

    def _append_batch(self, sheet_name: str, rows: List[list]) -> None:
        with self._lock:
            try:
                wb = self._load_workbook_safe()
                ws = wb[sheet_name]
                for row in rows:
                    ws.append(row)
                wb.save(self.path)
                self._maybe_trim()
            except Exception as exc:
                logger.warning(f"⚠️  ExcelWriter — échec écriture [{sheet_name}] : {exc}")

    def _maybe_trim(self) -> None:
        if self.path.stat().st_size <= self.max_bytes:
            return
        try:
            wb = self._load_workbook_safe()
            for ws in wb.worksheets:
                data_rows = ws.max_row - 1  # ligne 1 = en-tête
                to_delete = max(1, int(data_rows * _TRIM_FRACTION))
                for _ in range(to_delete):
                    ws.delete_rows(2)  # row 2 = la plus ancienne (après en-tête)
            wb.save(self.path)
            size_mb = self.path.stat().st_size / (1024 * 1024)
            logger.info(
                f"✂️  ExcelWriter — rotation effectuée "
                f"({_TRIM_FRACTION*100:.0f}% supprimé) → {size_mb:.1f} MB"
            )
        except Exception as exc:
            logger.warning(f"⚠️  ExcelWriter — échec rotation : {exc}")

    # ── Public async API ──────────────────────────────────────────────────

    async def write_metrics(
        self,
        vm_id: str,
        metrics: Dict[str, Any],
        timestamp: str,
        reliability: Any,
    ) -> None:
        row = [
            timestamp,
            vm_id,
            metrics.get("latency"),
            metrics.get("cpu_usage"),
            metrics.get("ram_usage"),
            reliability,
        ]
        await asyncio.to_thread(self._append_batch, "Métriques", [row])

    async def write_decision(self, payload: Dict[str, Any]) -> None:
        row = [
            payload.get("timestamp", datetime.now(timezone.utc).isoformat()),
            payload.get("decision"),
            payload.get("from_vm"),
            payload.get("to_vm"),
            payload.get("reason"),
            payload.get("cycle"),
        ]
        await asyncio.to_thread(self._append_batch, "Décisions", [row])

    async def write_slos(self, slos: List[Dict[str, Any]], timestamp: str) -> None:
        rows = [
            [
                timestamp,
                slo.get("metric"),
                slo.get("threshold"),
                slo.get("operator"),
                slo.get("weight"),
                slo.get("is_primary", False),
            ]
            for slo in slos
        ]
        if rows:
            await asyncio.to_thread(self._append_batch, "SLOs", rows)

    async def write_intent(self, intention: str, nb_slos: int) -> None:
        row = [datetime.now(timezone.utc).isoformat(), intention, nb_slos]
        await asyncio.to_thread(self._append_batch, "Intentions_LLM", [row])
