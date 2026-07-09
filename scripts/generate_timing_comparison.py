"""
scripts/generate_timing_comparison.py — Comparaison Autonomous vs Enhanced.

Lit les deux fichiers de mesures produits par le hub (TimingWriter) :
  - data/timings_autonomous.xlsx  (une ligne par cycle)
  - data/timings_enhanced.xlsx    (une ligne par intention)

et produit data/timings_comparison.xlsx : pour chaque étape instrumentée
commune aux deux modes (colonnes "ms" partagées entre les deux schémas de
shared/timing_writer.py), la moyenne / min / max par mode, plus une section
TOTAL séparée (cycle complet, et — enhanced uniquement — réception d'intention).

Script à lancer à la demande (pas de mise à jour automatique) :
    ./venv/Scripts/python.exe scripts/generate_timing_comparison.py
"""
import statistics
import sys

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from shared import config
from shared.timing_writer import (
    _AUTONOMOUS_COLS,
    _ENHANCED_COLS,
    _FILL_SUM,
    _HEADER_ROWS,
    _SHEET_AUTONOMOUS,
    _SHEET_ENHANCED,
    _SUM,
    _SVC_DARK,
    _SVC_MED,
)

_SVC_DISPLAY_OVERRIDE = {_SUM: "Résumé"}

_OUT_PATH = Path("data/timings_comparison.xlsx")
_SHEET_CMP = "Comparaison Autonomous-Enhanced"

_FONT_TITLE = Font(bold=True, color="1F3864", size=14)
_FONT_HEAD  = Font(bold=True, color="FFFFFF", size=10)
_FONT_SVC   = Font(bold=True, color="FFFFFF", size=10)
_ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_L    = Alignment(horizontal="left",   vertical="center")
_THIN       = Side(style="thin", color="9DB0CE")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FILL_HEAD  = PatternFill("solid", fgColor="1F3864")
_FILL_TOTAL = PatternFill("solid", fgColor="7B3F00")
_FILL_NA    = PatternFill("solid", fgColor="EAEAEA")

_HEADERS = [
    "Microservice", "Étape",
    "Autonomous n", "Autonomous moyenne (ms)", "Autonomous min (ms)", "Autonomous max (ms)",
    "Enhanced n", "Enhanced moyenne (ms)", "Enhanced min (ms)", "Enhanced max (ms)",
    "Δ moyenne (Enhanced − Autonomous)",
]


def _read_rows(path: Path, sheet_name: str, keys: List[str]) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = [dict(zip(keys, row)) for row in ws.iter_rows(min_row=_HEADER_ROWS + 1, values_only=True)]
    wb.close()
    return rows


def _stats(rows: List[Dict[str, Any]], key: str) -> Optional[Dict[str, float]]:
    vals = [r[key] for r in rows if isinstance(r.get(key), (int, float))]
    if not vals:
        return None
    return {
        "n":    len(vals),
        "mean": statistics.fmean(vals),
        "min":  min(vals),
        "max":  max(vals),
        "std":  statistics.pstdev(vals) if len(vals) > 1 else 0.0,
    }


def _shared_ms_columns() -> List[Tuple[str, str, str]]:
    """Colonnes 'ms' présentes dans les DEUX schémas (mêmes clé/service/libellé)."""
    auto_map = {c[0]: c for c in _AUTONOMOUS_COLS}
    enh_map  = {c[0]: c for c in _ENHANCED_COLS}
    shared: List[Tuple[str, str, str]] = []
    for key, svc, _sg, label, kind in _AUTONOMOUS_COLS:
        if kind != "ms" or svc is None:
            continue
        if key in enh_map and enh_map[key][4] == "ms":
            shared.append((key, _SVC_DISPLAY_OVERRIDE.get(svc, svc), label))
    return shared


def _write_row(ws, r: int, cells: List[Any], fills: Optional[Dict[int, PatternFill]] = None) -> None:
    fills = fills or {}
    for i, val in enumerate(cells, start=1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.border = _BORDER
        cell.alignment = _ALIGN_CTR if i > 2 else _ALIGN_L
        if i in fills:
            cell.fill = fills[i]
            cell.font = _FONT_HEAD


def _fmt(stat: Optional[Dict[str, float]], field: str) -> Any:
    if stat is None:
        return "—"
    return round(stat[field], 3) if field != "n" else stat["n"]


def main() -> None:
    auto_keys = [c[0] for c in _AUTONOMOUS_COLS]
    enh_keys  = [c[0] for c in _ENHANCED_COLS]

    auto_rows = _read_rows(Path(config.TIMING_EXCEL_AUTONOMOUS_PATH), _SHEET_AUTONOMOUS, auto_keys)
    enh_rows  = _read_rows(Path(config.TIMING_EXCEL_ENHANCED_PATH),   _SHEET_ENHANCED,   enh_keys)

    print(f"📂 Autonomous : {len(auto_rows)} ligne(s) lue(s)")
    print(f"📂 Enhanced   : {len(enh_rows)} ligne(s) lue(s)")

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_CMP

    ws.cell(row=1, column=1, value="Comparaison des temps d'exécution — Autonomous vs Enhanced").font = _FONT_TITLE
    ws.cell(row=2, column=1, value=(
        f"Généré le {datetime.now(timezone.utc).isoformat(timespec='seconds')} — "
        f"{len(auto_rows)} cycle(s) autonomous, {len(enh_rows)} intention(s) enhanced."
    ))

    r = 4
    _write_row(ws, r, _HEADERS, fills={i: _FILL_HEAD for i in range(1, len(_HEADERS) + 1)})
    r += 1

    last_svc: Optional[str] = None
    for key, svc, label in _shared_ms_columns():
        if svc != last_svc:
            fill = _FILL_SUM if svc == "Résumé" else _SVC_DARK.get(svc)
            cell = ws.cell(row=r, column=1, value=svc)
            if fill:
                cell.fill = fill
                cell.font = _FONT_SVC
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(_HEADERS))
            for c in range(1, len(_HEADERS) + 1):
                ws.cell(row=r, column=c).border = _BORDER
            r += 1
            last_svc = svc

        a = _stats(auto_rows, key)
        e = _stats(enh_rows, key)
        delta = round(e["mean"] - a["mean"], 3) if (a and e) else "—"
        _write_row(ws, r, [
            "", label,
            _fmt(a, "n"), _fmt(a, "mean"), _fmt(a, "min"), _fmt(a, "max"),
            _fmt(e, "n"), _fmt(e, "mean"), _fmt(e, "min"), _fmt(e, "max"),
            delta,
        ])
        r += 1

    # ── Section TOTAL — comparaison + colonnes enhanced-uniquement ──────────
    r += 1
    cell = ws.cell(row=r, column=1, value="TOTAL / RÉSUMÉ")
    cell.fill = _FILL_TOTAL
    cell.font = _FONT_SVC
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(_HEADERS))
    for c in range(1, len(_HEADERS) + 1):
        ws.cell(row=r, column=c).border = _BORDER
    r += 1

    a_total = _stats(auto_rows, "total")
    e_total = _stats(enh_rows, "cycle_total")
    delta_total = round(e_total["mean"] - a_total["mean"], 3) if (a_total and e_total) else "—"
    _write_row(ws, r, [
        "", "TOTAL cycle (autonomous: cycle complet — enhanced: cycle qui exécute la migration)",
        _fmt(a_total, "n"), _fmt(a_total, "mean"), _fmt(a_total, "min"), _fmt(a_total, "max"),
        _fmt(e_total, "n"), _fmt(e_total, "mean"), _fmt(e_total, "min"), _fmt(e_total, "max"),
        delta_total,
    ])
    r += 1

    e_recept = _stats(enh_rows, "intent_reception")
    _write_row(ws, r, [
        "", "Réception intention + LLM (Intent Manager) — enhanced uniquement, pas d'équivalent autonomous",
        "—", "—", "—", "—",
        _fmt(e_recept, "n"), _fmt(e_recept, "mean"), _fmt(e_recept, "min"), _fmt(e_recept, "max"),
        "—",
    ], fills={i: _FILL_NA for i in range(3, 7)})
    r += 1

    e_intent_total = _stats(enh_rows, "intention_total")
    _write_row(ws, r, [
        "", "TOTAL intention (réception → migration) — enhanced uniquement",
        "—", "—", "—", "—",
        _fmt(e_intent_total, "n"), _fmt(e_intent_total, "mean"), _fmt(e_intent_total, "min"), _fmt(e_intent_total, "max"),
        "—",
    ], fills={i: _FILL_NA for i in range(3, 7)})

    # ── Dimensions ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    for col in range(3, len(_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "C5"

    _OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_OUT_PATH)
    print(f"✅ Comparaison générée : {_OUT_PATH}")


if __name__ == "__main__":
    main()
