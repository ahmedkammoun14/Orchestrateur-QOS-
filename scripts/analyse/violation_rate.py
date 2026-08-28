"""
Taux de violation SLO (latence >= 28 ms) le long du tour.

LE CHIFFRE CENTRAL DU PAPIER. On croise deux sources independantes :
  - latences_session.csv : la trajectoire REELLE mesuree par le pont PiCar,
    verite terrain, une ligne par instant, une colonne de latence par VM ;
  - timings_autonomous_*.xlsx : quelle VM hebergeait le service a cet instant.

Pour chaque instant on lit la latence vers la VM qui servait REELLEMENT, et
on compte les instants ou elle depasse le seuil.

─────────────────────────────────────────────────────────────────────────
PIEGE 1 — LA COLONNE. Utiliser "VM active (source)" (= state.service_vm,
suivi fin), JAMAIS "VM hote (federation)" (= state.hosting_vm, issu de
kubectl). kubectl resout au NOEUD : edge1/edge1b/edge1c partagent
pop1-worker-1, il renvoie donc toujours la canonique edge1. En UC2 la
colonne federation vaut edge1 sur 100 % des cycles alors que le service
circulait sur 4 VMs. Se tromper de colonne donne 82,3 % au lieu de 53,7 %.

PIEGE 2 — L'ALIGNEMENT DES DATES (corrige le 17/08/2026). latences.csv ne
porte qu'une HEURE (HH:MM:SS), sans date, en heure locale du Pi (UTC+2) ;
les fichiers de temps sont en UTC avec date complete. L'ancienne version
ajoutait un jour des qu'un echantillon tombait plus de 10 min avant le
debut de la chronologie, pour gerer un passage de minuit. Mais quand le
PiCar a demarre AVANT l'orchestrateur — cas du run ABL du 17/08, 21 min
d'avance — cette regle projetait ces echantillons au lendemain, ou tous
les evenements de la chronologie leur etaient anterieurs : ils heritaient
donc de la DERNIERE VM du run. Environ 350 echantillons faussement
attribues.

Regle appliquee ici : on essaie les trois decalages (-1, 0, +1 jour) et on
ne retient que celui qui tombe DANS la fenetre couverte par la chronologie.
Aucun candidat valable => l'echantillon est ecarte, jamais devine.
"""
import csv
import datetime
import statistics
import openpyxl
from scipy.stats import mannwhitneyu

VMS_ORDER = ["edge1", "edge1b", "edge1c", "edge2", "edge2b", "edge2c",
             "cloud1", "cloud2"]
THRESHOLD_MS = 28.0
TZ_OFFSET_H = 2          # Pi en heure locale = UTC+2

# ── Les runs de la campagne ──────────────────────────────────────────
# (libelle, bras, dossier, fichiers de temps a fusionner)
# 4+4 complet au 18/08/2026 : campagne terminee, seuil du test de
# Mann-Whitney (p minimale 2/C(8,4)=0.029) desormais atteignable.
RUNS = [
    ("UC1  (13 aout)", "federe",   "data_UC1_federe",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("FED1 (17 aout)", "federe",   "data_FED_run1",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("FED2 (18 aout)", "federe",   "data_FED_run2",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("FED3 (18 aout)", "federe",   "data_FED_run3",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    # RUN 9 (21 aout) : HORS CAMPAGNE. Poids d'horizon [3,5,5,4,3,2,1] au lieu
    # de [7,6,5,4,3,2,1]. Ne pas l'inclure dans les moyennes ni dans le test de
    # Mann-Whitney : il tourne avec un code different des 8 autres.
    ("RUN9 (21 aout, horizon corrige)", "federe", "data_RUN9_horizon",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    # ── Journee LAAS du 27/08/2026 ────────────────────────────────────────
    # R1 = temoin du jour, configuration de la campagne (GRU seul, poids
    # [7,6,...,1]). R2/R3 = extrapolation lineaire + poids [1,2,...,7].
    # HORS CAMPAGNE eux aussi : ne pas les melanger aux moyennes des 4 runs
    # d'aout, ils servent a la comparaison temoin/ameliore du meme jour.
    ("R1   (27 aout, temoin)", "federe", "data_27-08_R1",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("R2   (27 aout, ameliore)", "federe", "data_27-08_R2",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("R3   (27 aout, ameliore)", "federe", "data_27-08_R3",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("R4   (27 aout, exponentiel)", "federe", "data_27-08_R4",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("R5   (27 aout, temoin 2)", "federe", "data_27-08_R5",
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"]),
    ("UC2  (14 aout)", "ablation", "data_UC2_ablation",
     ["timings_autonomous_provider1.xlsx"]),
    ("ABL1 (17 aout)", "ablation", "data_ABL_run1",
     ["timings_autonomous_provider1.xlsx"]),
    ("ABL2 (18 aout)", "ablation", "data_ABL_run2",
     ["timings_autonomous_provider1.xlsx"]),
    ("ABL3 (18 aout)", "ablation", "data_ABL_run3",
     ["timings_autonomous_provider1.xlsx"]),
]


def entetes(ws, n=3):
    """Aplatit un entete sur n lignes fusionnees en 'Groupe > Colonne'."""
    rows = list(ws.iter_rows(min_row=1, max_row=n, values_only=True))
    out, cur = [], [None] * n
    for c in range(ws.max_column):
        for r in range(n):
            v = rows[r][c] if c < len(rows[r]) else None
            if v is not None:
                cur[r] = v
                for k in range(r + 1, n):
                    cur[k] = None
        out.append(" > ".join(str(x) for x in cur if x))
    return out


def charger_chronologie(*chemins):
    """(instant UTC, VM servante) fusionnes et tries sur les N providers."""
    evenements = []
    for chemin in chemins:
        wb = openpyxl.load_workbook(chemin, read_only=True)
        ws = wb.worksheets[0]
        idx = {n: i for i, n in enumerate(entetes(ws))}
        i_ts = idx["Horodatage (UTC)"]
        i_vm = idx["VM active (source)"]          # cf. PIEGE 1
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r[0] is None or r[i_ts] is None or r[i_vm] is None:
                continue
            try:
                t = datetime.datetime.fromisoformat(str(r[i_ts])).replace(tzinfo=None)
            except ValueError:
                continue
            evenements.append((t, r[i_vm]))
        wb.close()
    evenements.sort(key=lambda e: e[0])
    dedup = []
    for t, vm in evenements:
        if dedup and dedup[-1] == (t, vm):
            continue
        dedup.append((t, vm))
    return dedup


def vm_servante_a(chronologie, t):
    """VM du dernier evenement anterieur ou egal a t."""
    best = None
    for et, vm in chronologie:
        if et <= t:
            best = vm
        else:
            break
    return best


def analyser(libelle, dossier, fichiers_temps):
    chrono = charger_chronologie(*[f"{dossier}/{f}" for f in fichiers_temps])
    debut, fin = chrono[0][0], chrono[-1][0]

    with open(f"{dossier}/latences_session.csv", newline="", encoding="utf-8") as f:
        lignes = list(csv.reader(f))[1:]

    total = violations = hors_fenetre = sans_latence = 0
    par_vm = {}

    for ligne in lignes:
        if len(ligne) != 19:
            continue
        h, m, s = map(int, ligne[0].split(":"))

        # cf. PIEGE 2 : on choisit le decalage de jour qui tombe DANS la
        # fenetre du run, sans jamais extrapoler hors de celle-ci.
        instant = None
        for delta_j in (0, 1, -1):
            cand = (datetime.datetime.combine(debut.date(), datetime.time(h, m, s))
                    - datetime.timedelta(hours=TZ_OFFSET_H)
                    + datetime.timedelta(days=delta_j))
            if debut <= cand <= fin:
                instant = cand
                break
        if instant is None:
            hors_fenetre += 1
            continue

        vm = vm_servante_a(chrono, instant)
        if vm not in VMS_ORDER:
            hors_fenetre += 1
            continue

        brut = ligne[3 + VMS_ORDER.index(vm)]
        if not brut:
            sans_latence += 1
            continue

        total += 1
        par_vm[vm] = par_vm.get(vm, 0) + 1
        if float(brut) >= THRESHOLD_MS:
            violations += 1

    pct = 100 * violations / total if total else float("nan")
    print(f"\n=== {libelle} — {dossier}")
    print(f"  fenetre      : {debut.strftime('%H:%M:%S')} -> {fin.strftime('%H:%M:%S')} UTC"
          f"  ({len(chrono)} evenements)")
    print(f"  echantillons : {total} retenus"
          f"  (hors fenetre : {hors_fenetre}, latence absente : {sans_latence})")
    print(f"  VMs servantes: {dict(sorted(par_vm.items(), key=lambda kv: -kv[1]))}")
    print(f"  >> VIOLATION (latence >= {THRESHOLD_MS:.0f} ms) : "
          f"{violations}/{total} = {pct:.1f} % du temps")
    return pct, total, violations


print("=" * 72)
print(f"  TAUX DE VIOLATION — seuil {THRESHOLD_MS:.0f} ms")
print("=" * 72)

# Runs qui ne font PAS partie de la campagne 4+4 : code ou configuration
# differents des huit runs de reference. Ils sont mesures et affiches, mais
# EXCLUS des moyennes et du test de Mann-Whitney — sans quoi on compare des
# executions qui ne tournent pas le meme systeme. Le commentaire de RUN9
# disait deja de l'exclure ; la synthese ne le faisait pas (corrige le
# 27/08/2026, apres avoir constate une moyenne federee a 26,0 % au lieu de
# 23,8 % et une p-valeur calculee sur un jeu melange).
HORS_CAMPAGNE = {
    "data_RUN9_horizon",   # poids d'horizon [3,5,5,4,3,2,1]
    "data_27-08_R1",       # temoin du 27/08
    "data_27-08_R2",       # extrapolation lineaire + poids [1..7]
    "data_27-08_R3",
    "data_27-08_R4",
    "data_27-08_R5",
}

resultats = {"federe": [], "ablation": []}
hors_campagne = []
for libelle, bras, dossier, fichiers in RUNS:
    pct, total, viol = analyser(libelle, dossier, fichiers)
    if dossier in HORS_CAMPAGNE:
        hors_campagne.append((libelle, pct, total, viol))
    else:
        resultats[bras].append((libelle, pct, total, viol))

if hors_campagne:
    print("\n" + "=" * 72)
    print("  HORS CAMPAGNE — mesures, mais exclus des moyennes et du test")
    print("=" * 72)
    for lib, pct, total, viol in hors_campagne:
        print(f"  {lib:<28} {pct:>5.1f} %   ({viol}/{total})")

print("\n" + "=" * 72)
print("  SYNTHESE PAR BRAS — campagne 4+4 uniquement")
print("=" * 72)
for bras in ("federe", "ablation"):
    valeurs = [p for _, p, _, _ in resultats[bras]]
    detail = "  ".join(f"{lib.split()[0]}={p:.1f}%" for lib, p, _, _ in resultats[bras])
    moyenne = statistics.mean(valeurs)
    # ecart-type d'echantillon : n-1 au denominateur. Sur 2 runs il vaut
    # simplement |a-b|/sqrt(2) ; il decrit la dispersion observee, il ne
    # fonde aucun test — 2+2 ne permet aucune p-valeur inferieure a 0,33.
    ecart = statistics.stdev(valeurs) if len(valeurs) > 1 else float("nan")
    print(f"\n  {bras.upper():<9} n={len(valeurs)}   {detail}")
    print(f"  {'':<9} moyenne {moyenne:.1f} %   ecart-type {ecart:.1f} pt")

f = [p for _, p, _, _ in resultats["federe"]]
a = [p for _, p, _, _ in resultats["ablation"]]
print(f"\n  Ecart ablation - federe : {statistics.mean(a) - statistics.mean(f):+.1f} points")

n, m = len(f), len(a)
if n >= 4 and m >= 4:
    stat, p = mannwhitneyu(a, f, alternative="two-sided", method="exact")
    print(f"\n  Test de Mann-Whitney (exact, bilateral), ablation vs federe :")
    print(f"    U = {stat:.1f}   p = {p:.4f}")
    from math import comb
    p_min = 2 / comb(n + m, n)
    print(f"    (p minimale atteignable avec n={n}, m={m} : {p_min:.4f})")
    if p < 0.05:
        print(f"    -> SIGNIFICATIF a 5% : les deux bras different bien.")
    else:
        print(f"    -> non significatif a 5% malgre le nombre de runs atteint.")
else:
    print(f"\n  ⚠ n={n}, m={m} par bras : Mann-Whitney non pertinent en dessous de 4+4")
    from math import comb
    p_min = 2 / comb(n + m, min(n, m))
    print(f"    (p minimale atteignable ici : {p_min:.4f}).")
    print("    Rapporter moyenne et etendue, jamais une p-valeur.\n")
