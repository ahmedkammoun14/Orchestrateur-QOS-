"""
Le taux de violation de l'orchestrateur est-il MAUVAIS, ou proche de
l'optimum atteignable sur ce banc ?

On compare, sur EXACTEMENT les memes instants que violation_rate.py :

  ORACLE        a chaque instant, la meilleure VM du pool. Borne inferieure
                absolue : aucun orchestrateur ne peut faire mieux, meme en
                connaissant le futur.
  NOTRE SYSTEME la VM reellement choisie (repris de violation_rate.py).
  MOYENNE       violation d'une VM tiree au hasard.
  PIRE VM       borne superieure.
  STATIQUE      chaque VM si le service y restait tout le tour, sans jamais
                migrer. C'est la reference qui compte : elle mesure ce que
                l'orchestration apporte par rapport a ne rien faire.

⚠ Les pourcentages "NOTRE SYSTEME" ci-dessous sont RECOPIES depuis
violation_rate.py. Ils doivent etre mis a jour ensemble. Les valeurs 60,7 %
et 82,3 % qui figuraient ici avant le 17/08/2026 etaient FAUSSES : elles
venaient de la colonne "VM hote (federation)" (resolution au noeud) au lieu
de "VM active (source)". Voir l'avertissement en tete de violation_rate.py.

⚠ Meme fenetre temporelle que violation_rate.py : les echantillons de
trajectoire anterieurs au demarrage de l'orchestrateur (21 min pour le run
ABL du 17/08) sont ecartes, sinon l'oracle serait calcule sur des instants
que le systeme n'a jamais eu a servir.
"""
import csv
import datetime
import statistics
import openpyxl

VMS = ["edge1", "edge1b", "edge1c", "edge2", "edge2b", "edge2c", "cloud1", "cloud2"]
POOL_P1 = ["edge1", "edge1b", "edge1c", "cloud1"]
THRESHOLD = 28.0
TZ_OFFSET_H = 2

# (libelle, dossier, pool, fichiers de temps, % mesure du systeme)
# Les % SYSTEME sont RECOPIES depuis violation_rate.py -- a mettre a jour
# ensemble (voir avertissement en tete de ce fichier).
RUNS = [
    ("UC1  (13 aout) — FEDERE, 8 VMs", "data_UC1_federe", VMS,
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"], 19.9),
    ("FED1 (17 aout) — FEDERE, 8 VMs", "data_FED_run1", VMS,
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"], 25.9),
    ("FED2 (18 aout) — FEDERE, 8 VMs", "data_FED_run2", VMS,
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"], 29.3),
    ("FED3 (18 aout) — FEDERE, 8 VMs", "data_FED_run3", VMS,
     ["timings_autonomous_provider1.xlsx", "timings_autonomous_provider2.xlsx"], 20.2),
    ("UC2  (14 aout) — ABLATION, 4 VMs", "data_UC2_ablation", POOL_P1,
     ["timings_autonomous_provider1.xlsx"], 53.7),
    ("ABL1 (17 aout) — ABLATION, 4 VMs", "data_ABL_run1", POOL_P1,
     ["timings_autonomous_provider1.xlsx"], 52.7),
    ("ABL2 (18 aout) — ABLATION, 4 VMs", "data_ABL_run2", POOL_P1,
     ["timings_autonomous_provider1.xlsx"], 56.6),
    ("ABL3 (18 aout) — ABLATION, 4 VMs", "data_ABL_run3", POOL_P1,
     ["timings_autonomous_provider1.xlsx"], 56.3),
]


def fenetre_du_run(dossier, fichiers):
    """Premier et dernier horodatage UTC couverts par les fichiers de temps."""
    bornes = []
    for f in fichiers:
        wb = openpyxl.load_workbook(f"{dossier}/{f}", read_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        cur, entetes = [None] * 3, []
        for c in range(ws.max_column):
            for r in range(3):
                v = rows[r][c] if c < len(rows[r]) else None
                if v is not None:
                    cur[r] = v
                    for k in range(r + 1, 3):
                        cur[k] = None
            entetes.append(" > ".join(str(x) for x in cur if x))
        i_ts = {n: i for i, n in enumerate(entetes)}["Horodatage (UTC)"]
        ts = []
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r[0] is None or r[i_ts] is None:
                continue
            try:
                ts.append(datetime.datetime.fromisoformat(str(r[i_ts])).replace(tzinfo=None))
            except ValueError:
                continue
        wb.close()
        if ts:
            bornes += [min(ts), max(ts)]
    return min(bornes), max(bornes)


def charger(dossier, debut, fin):
    """Latences par VM, restreintes a la fenetre du run."""
    with open(f"{dossier}/latences_session.csv", newline="", encoding="utf-8") as f:
        lignes = list(csv.reader(f))[1:]
    out = []
    for r in lignes:
        if len(r) != 19:
            continue
        h, m, s = map(int, r[0].split(":"))
        instant = None
        for delta_j in (0, 1, -1):
            cand = (datetime.datetime.combine(debut.date(), datetime.time(h, m, s))
                    - datetime.timedelta(hours=TZ_OFFSET_H)
                    + datetime.timedelta(days=delta_j))
            if debut <= cand <= fin:
                instant = cand
                break
        if instant is None:
            continue
        vals = {vm: float(r[3 + i]) for i, vm in enumerate(VMS) if r[3 + i]}
        if vals:
            out.append(vals)
    return out


def analyser(libelle, dossier, pool, fichiers, pct_systeme):
    debut, fin = fenetre_du_run(dossier, fichiers)
    data = charger(dossier, debut, fin)
    n = len(data)
    if not n:
        print(f"{libelle} : aucune donnee")
        return None

    oracle = pire = 0
    par_vm = {vm: 0 for vm in pool}
    taux_moyens = []

    for vals in data:
        dispo = {vm: vals[vm] for vm in pool if vm in vals}
        if not dispo:
            continue
        if min(dispo.values()) >= THRESHOLD:
            oracle += 1
        if max(dispo.values()) >= THRESHOLD:
            pire += 1
        for vm, v in dispo.items():
            if v >= THRESHOLD:
                par_vm[vm] += 1
        taux_moyens.append(sum(1 for v in dispo.values() if v >= THRESHOLD) / len(dispo))

    pct_oracle = 100 * oracle / n
    statiques = {vm: 100 * par_vm[vm] / n for vm in pool}
    meilleur_statique = min(statiques.values())

    print(f"\n=== {libelle}")
    print(f"  instants retenus : {n}   (fenetre {debut.strftime('%H:%M')} -> {fin.strftime('%H:%M')} UTC)")
    print(f"  ORACLE   (meilleure VM a chaque instant) : {pct_oracle:5.1f} %   <- borne min")
    print(f"  SYSTEME  (mesure)                        : {pct_systeme:5.1f} %")
    print(f"  MOYENNE  (VM au hasard)                  : {100*statistics.mean(taux_moyens):5.1f} %")
    print(f"  PIRE VM                                  : {100*pire/n:5.1f} %   <- borne max")
    print(f"  --- placement STATIQUE, aucune migration ---")
    for vm in sorted(pool, key=lambda v: statiques[v]):
        marque = "  <- meilleur statique" if statiques[vm] == meilleur_statique else ""
        print(f"      rester sur {vm:8s} : {statiques[vm]:5.1f} %{marque}")

    # Part du gain atteignable effectivement capturee :
    #   du meilleur statique (ne rien faire) jusqu'a l'oracle (ideal).
    marge = meilleur_statique - pct_oracle
    capture = 100 * (meilleur_statique - pct_systeme) / marge if marge > 0 else float("nan")
    print(f"  >> ecart a l'oracle        : {pct_systeme - pct_oracle:+.1f} points")
    print(f"  >> gain sur meilleur statique : {meilleur_statique - pct_systeme:+.1f} points"
          f"  ({capture:.0f} % du gain atteignable)")
    return {"oracle": pct_oracle, "systeme": pct_systeme,
            "statique": meilleur_statique, "capture": capture}


print("=" * 74)
print(f"  BORNES DE COMPARAISON — seuil {THRESHOLD:.0f} ms")
print("=" * 74)

res = {}
for libelle, dossier, pool, fichiers, pct in RUNS:
    bras = "federe" if "FEDERE" in libelle else "ablation"
    r = analyser(libelle, dossier, pool, fichiers, pct)
    if r:
        res.setdefault(bras, []).append(r)

print("\n" + "=" * 74)
print("  SYNTHESE")
print("=" * 74)
for bras in ("federe", "ablation"):
    if bras not in res:
        continue
    lot = res[bras]
    moy = lambda k: statistics.mean(x[k] for x in lot)
    print(f"\n  {bras.upper():<9} n={len(lot)}")
    print(f"    oracle            : {moy('oracle'):5.1f} %")
    print(f"    meilleur statique : {moy('statique'):5.1f} %")
    print(f"    systeme mesure    : {moy('systeme'):5.1f} %")
    print(f"    gain capture      : {moy('capture'):5.0f} % du gain atteignable")
print()
