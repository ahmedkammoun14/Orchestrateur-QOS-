"""
MAE/RMSE des 3 modeles ML, recalcules depuis les feuilles Predictions
des archives reelles (UC1, UC2, UC4-LAAS).

Methode : predicted[cycle=N] est une prevision a horizon 1 (le modele
retourne un horizon de 7, on ne loggue que le premier pas) pour le
PROCHAIN cycle de CETTE VM -- jamais le cycle courant. On compare donc
predicted[N] au measured du cycle suivant REELEMENT observe pour la
meme VM/metrique (pas N+1 suppose : des cycles peuvent manquer).
"""
import openpyxl, sys, math, collections

SOURCES = [
    ("UC1  (federe, 13 aout)",   "data_UC1_federe/qos_history_provider1.xlsx"),
    ("UC1  (federe, 13 aout)",   "data_UC1_federe/qos_history_provider2.xlsx"),
    ("FED1 (federe, 17 aout)",   "data_FED_run1/qos_history_provider1.xlsx"),
    ("FED1 (federe, 17 aout)",   "data_FED_run1/qos_history_provider2.xlsx"),
    ("FED2 (federe, 18 aout)",   "data_FED_run2/qos_history_provider1.xlsx"),
    ("FED2 (federe, 18 aout)",   "data_FED_run2/qos_history_provider2.xlsx"),
    ("FED3 (federe, 18 aout)",   "data_FED_run3/qos_history_provider1.xlsx"),
    ("FED3 (federe, 18 aout)",   "data_FED_run3/qos_history_provider2.xlsx"),
    ("UC2  (ablation, 14 aout)", "data_UC2_ablation/qos_history_provider1.xlsx"),
    ("ABL1 (ablation, 17 aout)", "data_ABL_run1/qos_history_provider1.xlsx"),
    ("ABL2 (ablation, 18 aout)", "data_ABL_run2/qos_history_provider1.xlsx"),
    ("ABL3 (ablation, 18 aout)", "data_ABL_run3/qos_history_provider1.xlsx"),
]

# metric -> {(vm, cycle): measured}, et liste (cycle, predicted) par vm
by_metric = collections.defaultdict(lambda: collections.defaultdict(dict))
pred_rows = collections.defaultdict(lambda: collections.defaultdict(list))

n_files = 0
for label, path in SOURCES:
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except FileNotFoundError:
        print(f"(absent, ignore) {path}")
        continue
    if "Prédictions" not in wb.sheetnames:
        print(f"(pas de feuille Predictions) {path}")
        wb.close()
        continue
    n_files += 1
    for ts, cyc, vm, metric, measured, predicted, model in \
            wb["Prédictions"].iter_rows(min_row=2, values_only=True):
        if ts is None or cyc is None or vm is None:
            continue
        if measured is not None:
            by_metric[metric][vm][cyc] = float(measured)
        if predicted is not None:
            pred_rows[metric][vm].append((cyc, float(predicted), model))
    wb.close()

print(f"\n{n_files} fichier(s) source charge(s)\n")

UNITS = {"latency": "ms", "cpu_usage": "%", "ram_usage": "%"}

for model_filter, model_label in (
    ("GRU", "Niveau 1 -- GRU (le modele)"),
    ("point_model", "Niveau 2 -- point_model (repli)"),
    ("last_value_fallback", "Niveau 3 -- persistance naive (pas une prediction)"),
):
    print(f"\n=== {model_label} ===")
    print(f"{'Métrique':12s} {'n paires':>9s} {'MAE':>10s} {'RMSE':>10s}  "
          f"{'unité':6s}  {'moy(mesuré)':>12s}  {'RMSE/moy':>9s}")
    print("-" * 78)

    for metric in ("latency", "cpu_usage", "ram_usage"):
        errors = []
        measured_vals = []
        for vm, preds in pred_rows[metric].items():
            preds_sorted = sorted(preds, key=lambda x: x[0])
            measured_for_vm = by_metric[metric][vm]
            available_cycles = sorted(measured_for_vm.keys())
            for cyc, pval, model in preds_sorted:
                if model != model_filter:
                    continue
                # cycle suivant REELEMENT observe pour cette VM (pas cyc+1 suppose)
                nxt = next((c for c in available_cycles if c > cyc), None)
                if nxt is None:
                    continue
                mval = measured_for_vm[nxt]
                errors.append(pval - mval)
                measured_vals.append(mval)

        if not errors:
            print(f"{metric:12s}  aucune paire alignable")
            continue

        n = len(errors)
        mae = sum(abs(e) for e in errors) / n
        rmse = math.sqrt(sum(e * e for e in errors) / n)
        mean_m = sum(measured_vals) / n
        ratio = rmse / mean_m if mean_m else float("nan")
        print(f"{metric:12s} {n:9d} {mae:10.3f} {rmse:10.3f}  "
              f"{UNITS.get(metric,''):6s}  {mean_m:12.3f}  {ratio:9.3f}")

total = sum(len(v) for m in pred_rows.values() for v in m.values())
gru = sum(1 for m in pred_rows.values() for v in m.values() for _,_,mo in v if mo=="GRU")
print(f"\nTaux de disponibilité du modèle niveau 1 (GRU) : "
      f"{gru}/{total} = {100*gru/total:.1f}%")
