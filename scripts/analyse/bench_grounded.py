"""
Banc AVANT/APRES de l'ancrage des seuils (INTENT_GROUNDED_THRESHOLDS).

POURQUOI HORS-LIGNE. Traduire une intention est une fonction pure de
(prompt, phrase) : cela ne mobilise ni le PiCar, ni les VMs, ni les
orchestrateurs. On peut donc mesurer l'effet du correctif sans refaire un
run — et surtout en COMPARAISON APPARIEE, les deux conditions partageant
exactement les memes percentiles. Un run complet ferait varier la
trajectoire, donc les percentiles, donc les seuils : on ne saurait plus si
l'ecart vient du prompt ou du trajet.

CE QU'IL MESURE / CE QU'IL NE MESURE PAS. Il mesure la sortie du LLM. Il
court-circuite _normalize_and_validate et le SLOMerger, donc il ne prouve
pas que la chaine complete fonctionne : cela demande un passage en
conditions reelles, court, apres la campagne.

Usage :
    python scripts/analyse/bench_grounded.py            # 12 phrases x 2
    python scripts/analyse/bench_grounded.py --repeat 3 # + test de stabilite
"""
import argparse
import asyncio
import importlib
import os
import statistics
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

RACINE = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(RACINE))
os.chdir(RACINE)

# Les 12 phrases d'UC5, dans l'ordre du run du 17/08/2026.
INTENTIONS = [
    ("#03", "Reduce latency as much as possible, I'm watching a live event"),
    ("#17", "Always encrypt my stream end-to-end, even if it increases latency"),
    ("#05", "Pre-load the next episode before it starts so there is no gap between episodes"),
    ("#02", "I always want at least 1080p. If the network cannot support it, pause and wait"),
    ("#01", "I want the video to never buffer, even if that means reducing quality"),
    ("#06", "Optimize my connection at every step of my commute from Toulouse to Bordeaux"),
    ("#12", "Minimize the carbon footprint of my stream. Use green data centres when possible"),
    ("#16", "Do not log or retain any metadata about my viewing sessions"),
    ("#18", "Alert me if my connection is being intercepted or downgraded"),
    ("#07", "When the car stops in traffic, use that opportunity to download and buffer more content"),
    ("#04", "Switch to audio-only mode when I'm in a tunnel or weak signal area"),
    ("#11", "Alert me and downgrade quality before I exceed my monthly data plan"),
]

SOURCE_LATENCES = [
    "data_UC5_intentions/qos_history_provider1.xlsx",
    "data_UC5_intentions/qos_history_provider2.xlsx",
]
SEUIL_AUTONOME = 28.0


def percentile(tries, p):
    if len(tries) == 1:
        return tries[0]
    k = (len(tries) - 1) * p / 100.0
    lo, hi = int(k), min(int(k) + 1, len(tries) - 1)
    return tries[lo] + (tries[hi] - tries[lo]) * (k - lo)


def percentiles_mesures():
    """Distribution des latences MESUREES pendant le run UC5 (archives)."""
    vals = []
    for chemin in SOURCE_LATENCES:
        try:
            wb = openpyxl.load_workbook(chemin, read_only=True)
        except FileNotFoundError:
            continue
        if "Prédictions" in wb.sheetnames:
            for ligne in wb["Prédictions"].iter_rows(min_row=2, values_only=True):
                # (timestamp, cycle, vm, metrique, mesure, prediction, modele)
                if len(ligne) >= 5 and ligne[3] == "latency" and ligne[4] is not None:
                    vals.append(float(ligne[4]))
        wb.close()
    if not vals:
        raise SystemExit("Aucune latence mesuree trouvee dans data_UC5_intentions/")
    vals.sort()
    return {f"p{p}": round(percentile(vals, p), 1) for p in (10, 25, 50, 75, 90)} | {"n": len(vals)}


def charger(actif: bool):
    """Recharge config puis llm_handler avec le drapeau demande."""
    os.environ["INTENT_GROUNDED_THRESHOLDS"] = "true" if actif else "false"
    for nom in [k for k in list(sys.modules) if k == "shared.config" or "llm_handler" in k]:
        del sys.modules[nom]
    conf = importlib.import_module("shared.config")
    assert conf.INTENT_GROUNDED_THRESHOLDS is actif
    return importlib.import_module("services.intent_manager.llm_handler")


def latence_de(slos):
    for s in slos or []:
        if s.get("metric") == "latency":
            return s.get("threshold"), s.get("weight")
    return None, None


async def traduire(module, phrase, pct):
    """Un appel REEL au LLM primaire, avec le prompt du mode courant."""
    h = module.LLMHandler()

    async def faux_percentiles(_self=None):
        return pct

    h._observed_latency_percentiles = faux_percentiles
    try:
        return await h._level1_llm(phrase, {"active_slos": [], "last_intention": None})
    except Exception as exc:                      # panne reseau, 404, timeout
        return ("ERREUR", str(exc)[:60])


def resume(res):
    """(texte lisible, seuil de latence) a partir du retour de _level1_llm."""
    if res is None:
        return "rejetee", None
    strategie, slos = res
    if strategie == "ERREUR":
        return f"ERREUR {slos}", None
    if not slos:
        return "rejetee", None
    seuil, poids = latence_de(slos)
    morceaux = []
    for s in slos:
        morceaux.append(f"{s.get('metric')}{s.get('operator','')}{s.get('threshold')}")
    return " ; ".join(morceaux), seuil


async def principal(repetitions):
    pct = percentiles_mesures()
    print(f"\nPercentiles des latences MESUREES pendant UC5 ({pct['n']} points)")
    print(f"  P10 {pct['p10']}   P25 {pct['p25']}   P50 {pct['p50']}   "
          f"P75 {pct['p75']}   P90 {pct['p90']}  ms")
    print(f"  seuil du mode autonome : {SEUIL_AUTONOME} ms\n")

    mod_avant = charger(False)
    mod_apres = charger(True)

    lignes_md = [
        "# Ancrage des seuils — comparaison avant / après",
        "",
        f"Généré le {datetime.now():%Y-%m-%d %H:%M}. Appels réels au LLM primaire.",
        "",
        f"Percentiles des latences mesurées pendant UC5 ({pct['n']} points) : "
        f"P10 {pct['p10']} · P25 {pct['p25']} · P50 {pct['p50']} · "
        f"P75 {pct['p75']} · P90 {pct['p90']} ms.",
        f"Seuil du mode autonome : **{SEUIL_AUTONOME} ms**.",
        "",
        "| # | Intention | Avant (latence) | Après (latence) |",
        "|---|---|---|---|",
    ]

    avants, apres_ = [], []
    print(f"{'#':<5}{'avant':>10}{'apres':>10}   intention")
    print("-" * 78)

    for num, phrase in INTENTIONS:
        r_av = await traduire(mod_avant, phrase, pct)
        r_ap = await traduire(mod_apres, phrase, pct)
        txt_av, seuil_av = resume(r_av)
        txt_ap, seuil_ap = resume(r_ap)

        aff_av = f"{seuil_av:.0f}" if seuil_av else ("rejet" if txt_av == "rejetee" else "--")
        aff_ap = f"{seuil_ap:.0f}" if seuil_ap else ("rejet" if txt_ap == "rejetee" else "--")
        print(f"{num:<5}{aff_av:>10}{aff_ap:>10}   {phrase[:44]}")

        if seuil_av:
            avants.append(seuil_av)
        if seuil_ap:
            apres_.append(seuil_ap)

        lignes_md.append(
            f"| {num} | `{phrase[:56]}` | {txt_av} | {txt_ap} |"
        )

    print("-" * 78)
    lignes_md += ["", "## Lecture", ""]

    if avants and apres_:
        med_av, med_ap = statistics.median(avants), statistics.median(apres_)
        sous_av = sum(1 for v in avants if v < SEUIL_AUTONOME)
        sous_ap = sum(1 for v in apres_ if v < SEUIL_AUTONOME)
        for txt in (
            f"seuils de latence produits : {len(avants)} avant, {len(apres_)} apres",
            f"mediane : {med_av:.0f} ms  ->  {med_ap:.0f} ms",
            f"plus STRICTS que le defaut autonome ({SEUIL_AUTONOME:.0f} ms) : "
            f"{sous_av}/{len(avants)}  ->  {sous_ap}/{len(apres_)}",
        ):
            print("  " + txt)
            lignes_md.append(f"- {txt}")

    if repetitions > 1:
        print(f"\nStabilite : {repetitions} tirages de #03 en mode ancre")
        lignes_md += ["", f"## Stabilité ({repetitions} tirages de #03, mode ancré)", ""]
        vus = []
        for _ in range(repetitions):
            r = await traduire(mod_apres, INTENTIONS[0][1], pct)
            _, seuil = resume(r)
            vus.append(seuil)
        print(f"  seuils obtenus : {vus}")
        lignes_md.append(f"- seuils obtenus : `{vus}`")
        lignes_md.append(
            "- " + ("identiques — traduction deterministe"
                    if len(set(vus)) == 1 else
                    "**divergents** — la traduction n'est pas deterministe")
        )

    sortie = RACINE / "logs" / f"bench_grounded_{datetime.now():%Y%m%d_%H%M}.md"
    sortie.parent.mkdir(exist_ok=True)
    sortie.write_text("\n".join(lignes_md) + "\n", encoding="utf-8")
    print(f"\nRapport : {sortie}\n")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--repeat", type=int, default=1,
                    help="tirages supplementaires pour tester le determinisme")
    asyncio.run(principal(ap.parse_args().repeat))
